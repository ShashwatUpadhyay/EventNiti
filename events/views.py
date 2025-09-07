from django.shortcuts import render,redirect,HttpResponse, get_object_or_404
from . import models
from django.contrib import messages
from django.contrib.auth.models import User
from account.models import UserExtra
from blog.models import Blog
from django.http import HttpResponseRedirect
from base.views import is_head, is_member, is_student, is_teacher
from certificate.views import generate_qr_code_base64
from django.contrib.auth.decorators import login_required
from ppuu import settings
from django.http import JsonResponse
from datetime import datetime
import pytz
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
from django.core.paginator import Paginator
india_timezone = pytz.timezone('Asia/Kolkata')
import logging 
from django.contrib.admin.views.decorators import staff_member_required
logger = logging.getLogger(__name__)

def is_cordinator(request,event):
    return event.cordinators.filter(user=request.user).exists()

def is_event_host(request,event): 
    return event.organized_by == request.user or request.user.is_superuser
    
def registered_in_event(request,event):
    return event.participant.filter(user=request.user).exists()

def events(request):
    event = models.Event.objects.filter(event_open=True,event_over = False,status='approved').order_by('start_date','-registration_open')
    over_event = models.Event.objects.filter(event_open=True,event_over = True,status='approved').order_by('-start_date')
    
    p = Paginator(event,3)
    page = request.GET.get('page')
    event = p.get_page(page)
    
    return render(request, 'events.html',{'event':event,'over_event':over_event})

@login_required(login_url='login')
def event(request, slug):
    try:
        event = models.Event.objects.get(slug= slug,status='approved')
    except Exception as e:
        return redirect('home')
    return render(request, 'event.html',{'event':event,'registered': registered_in_event(request,event)})

@login_required(login_url='login')
def eventregister(request, slug):
    user_obj = None
    event=None
    logger.info(f'User {request.user.username} is attempting to register for event with slug: {slug}')
    try:
        user_obj = UserExtra.objects.get(user=request.user)
    except Exception as e:
        print(e)
    try:
        event = models.Event.objects.get(slug= slug,status='approved')
        msg = False
        if  event.event_over:
            messages.error(request, "Event is Over!")
            msg = True
        if not event.registration_open:
            messages.error(request, "Registration is Closed!")
            msg = True
        if models.EventSubmission.objects.filter(user=request.user, event=event).exists():
            messages.error(request, "Rejected!")
            msg = True
        if msg:
            return redirect('event', slug =slug)
    except Exception as e:
        logger.error(f'Error fetching event with slug {slug}: {e}')
        messages.error(request, f"Error: {e}   ")  
        return redirect('home')
    
    if event.limit and event.count >= event.limit:
        event.registration_open = False
        event.text_status = 'Registration Full'
        event.save()    
        messages.error(request, "Registration is Full!")
        return redirect('event', slug =slug)
    
    if request.method == "POST":
        uuid = request.POST.get('uuid')
        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        course = request.POST.get('course')
        section = request.POST.get('section')
        year = request.POST.get('year')
        logger.debug(f'Form data received: uuid={uuid}, full_name={full_name}, email={email}, course={course}, section={section}, year={year}')
        if event.price > 0:
            submission , _ = models.TemporaryEventSubmission.objects.get_or_create(uu_id=uuid,full_name = full_name,year=year,email = email,user=request.user,course = course,section = section,event = event)
            logger.info(f'Temporary submission created for user {request.user.username} for event {event.title}')
            return redirect('payment', slug=event.slug, token=submission.uid)
        else:
            models.EventSubmission.objects.create(uu_id=uuid,full_name = full_name,year=year,email = email,user=request.user,course = course,section = section,event = event)
            logger.info(f'Free event submission created for user {full_name} for event {event.title}')

        logger.info(f'User {request.user.username} successfully registered for event {event.title}')
        messages.success(request,f"Submission Successful in {event.title} event")
        return redirect('events')

    return render(request, 'eventregister.html',{'event':event, 'user_obj':user_obj})

@login_required(login_url='login')
def create_event(request):
    if not request.user.groups.filter(name__in=['HOST']).exists() and not request.user.is_staff:
        return redirect('events')
    
    if request.method == "POST":
        title = request.POST.get('title')
        description = request.POST.get('description')
        start_date = request.POST.get('start_date')
        location = request.POST.get('location')
        end_date = request.POST.get('end_date')
        last_date_of_registration = request.POST.get('last_date_of_registration')
        poster = request.FILES.get('poster')
        limit = request.POST.get('limit')
        price = request.POST.get('price')
        offers_certification = request.POST.get('offers_certification')
        notify = request.POST.get('notify')
        registration_open = request.POST.get('registration_open')
        event_open = request.POST.get('event_open')
        print(title,description,start_date,location,end_date,last_date_of_registration,poster,limit,price,offers_certification,notify,registration_open,event_open)
        try:
            models.Event.objects.create(
                title=title, 
                description=description, 
                organized_by=request.user,
                start_date=start_date, 
                location=location, 
                end_date=end_date, 
                last_date_of_registration=last_date_of_registration, 
                poster=poster, 
                limit=limit, 
                price=price, 
                offers_certification=True if offers_certification=='on' else False, 
                notify=True if notify=='on' else False, 
                registration_open=True if registration_open=='on' else False, 
                event_open=True if event_open=='on' else False,
                )
            messages.success(request, "Event sent to admin for approval")
        except Exception as e:
            print(e)
            messages.error(request, f"Error: {e}   ")  
            return redirect('create_event')
    return render(request, 'events/create_event.html')

def eventTicket(request, uid):
    ticket = None
    submisson = None
    try:
        ticket = models.EventTicket.objects.get(uid=uid)
    except:
        return HttpResponse("invalid Ticket ID!")
    
    return render(request,'eventticket.html',{'ticket':ticket, 'qrCode':generate_qr_code_base64(f'{settings.DOMAIN_NAME}events/attendence/{ticket.submission_uid}/')})
    
@login_required(login_url='login')
def myTicket(request):
    ticket = models.EventTicket.objects.filter(user = request.user).order_by('-created_at')
    return render(request , 'tickets.html', {'ticket':ticket})
    
@login_required(login_url='login')
def takeSudentAttendence(request, submissionid):
    submission = get_object_or_404(models.EventSubmission , uid=submissionid)    
    if is_cordinator(request,submission.event) or is_event_host(request,submission.event) or request.user.is_staff:
        pass
    else:
        messages.error(request , 'Access denied')
        return redirect('events')
    
    if models.EventTicket.objects.get(submission_uid=submissionid).restricted:
        return render(request, 'eventattendence.html', {'submission': False, 'msg': 'Student is Restricted!', 'student': submission.full_name})

    if not submission.allowed:
        return render(request, 'eventattendence.html', {'submission': False, 'msg': 'Candidate is not allowed!', 'student': submission.full_name})
    
    if submission.attendence == 'Present':
        return render(request, 'eventattendence.html', {'submission': False, 'msg': 'Attendence Already Marked', 'student': submission.full_name})
    
    if submission.attendence == 'Absent':
        current_time_india = datetime.now(india_timezone)
        submission.attendence = 'Present'
        submission.attendence_taken_by = f'{request.user.get_full_name()} - {current_time_india.strftime('%Y-%m-%d %I:%M:%S %p')}'
        submission.save()
        
    return render(request, 'eventattendence.html', {'submission': True, 'msg': 'Attendence sucessfully Marked!! ', 'student': submission.full_name})  


@login_required(login_url='login')
def teacherEventList(request):
    if is_student(request.user):
        return redirect('events')
    
    event = models.Event.objects.filter(event_over=False, event_open=True).order_by('start_date')
    return render(request, 'eventsteacher.html',{'event':event})


@login_required(login_url='login')
def teacherEvent(request, slug):
    event = get_object_or_404(models.Event , slug=slug)
    if is_cordinator(request,event) or request.user.is_staff or is_event_host(request,event):
        return render(request, 'vieweventteacher.html',{'event':event , 'is_host':is_event_host(request,event)})
    return redirect('events')

@login_required(login_url='login')
def registeredStudentList(request, slug):    
    students = models.EventSubmission.objects.filter(event__slug=slug).order_by('full_name')
    event = get_object_or_404(models.Event , slug=slug)
    if is_cordinator(request,event) or request.user.is_staff or is_event_host(request,event):
        return render(request , 'registeredstudent.html', {'students':students, 'event' : event}) 
    return redirect('events')

@login_required(login_url='login')
def registeredStudentListExcel(request, slug):
    india_timezone = pytz.timezone("Asia/Kolkata")
    students = models.EventSubmission.objects.filter(event__slug=slug).order_by('full_name')
    event = get_object_or_404(models.Event, slug=slug)

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header row
    headers = ['Name', 'UUID', 'Course', 'Section', 'Year', 'Attendance', 'Attendance Taken By']
    ws.append(headers)

    # Color fills
    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")  # Present
    red_fill = PatternFill(start_color="FFCC0000", end_color="FFCC0000", fill_type="solid")    # Absent

    # Write student data
    for student in students:
        row_data = [
            student.user.get_full_name(),
            student.user.user_extra.uu_id,
            student.user.user_extra.course,
            student.user.user_extra.section,
            student.user.user_extra.year,
            student.attendence,
            student.attendence_taken_by,
        ]
        ws.append(row_data)

        # Get the last row index
        current_row = ws.max_row
        print(current_row)
        # Apply color to "Attendance" column (index 6, Excel is 1-based)
        attendance_cell = ws.cell(row=current_row, column=6)
        print(attendance_cell)
        if str(student.attendence).strip().lower() == 'present':
            attendance_cell.fill = green_fill
        elif str(student.attendence).strip().lower() == 'absent':
            attendance_cell.fill = red_fill

    # Save to memory stream instead of file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Response as XLSX
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{event.slug}_{datetime.now(india_timezone).strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
    
    # return redirect('events')

@login_required(login_url='login')
def registeredStudentListAjax(request, slug):
    event = models.Event.objects.get(slug=slug)
    if is_cordinator(request,event) or is_event_host(request,event) or request.user.is_staff:
        pass
    else:
        messages.error(request, "Access Denied!")
        return redirect('home')
    students = models.EventSubmission.objects.filter(event__slug=slug).order_by('full_name')
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':  # Check if it's an AJAX request
        student_data = []
        for student in students:
            student_data.append({
                "uu_id": student.uu_id,
                "full_name": student.full_name,
                "email": student.email,
                "course": student.course,
                "section": student.section,
                "year": student.year,
                "attendence": student.attendence,
                "attendence_taken_by": student.attendence_taken_by or "-",
                "allowed": "Yes" if student.allowed else "No"
            })
        return JsonResponse({"students": student_data})
    
    return render(request, 'registeredstudent.html', {'students': students, 'event': event})

def eventResult(request, slug):
    event =  None
    result =  None
    submisson =  None
    blog =  None
    try:
        event = models.Event.objects.get(slug=slug)
        submisson = models.EventSubmission.objects.filter(event = event)
        result = models.EventResult.objects.get(event = event)
        blogs = Blog.objects.filter(related_event = event)
    except Exception as e:
        messages.error(request,"No Result Announcement")
        return redirect('home')
    
    return render(request , 'resultofevent.html',{'event':event, 'result':result,'submisson':submisson,'blogs':blogs})

@login_required(login_url='login')
def my_coordinated_events(request):
    # coordinated_events = models.EventCordinator.objects.filter(user = request.user)
    coordinated_events = request.user.cordinator.all
    print(coordinated_events)
    return render(request, 'events/coordinated_events.html', {
        'events': coordinated_events
    })
    
@login_required(login_url='login')
def my_hosted_events(request):
    host_events = request.user.host.all
    return render(request, 'events/hosted_events.html', {
        'events': host_events
    })


@login_required(login_url='login')
def edit_event(request, slug):
    event = get_object_or_404(models.Event , slug=slug)
    if not is_event_host(request,event):
        return redirect('events')
    
    if request.method == "POST":
        title = request.POST.get('title')
        description = request.POST.get('description')
        location = request.POST.get('location')
        poster = request.FILES.get('poster')
        start_date = request.POST.get('start_date')
        start_time = request.POST.get('start_time')  
        registration_open = request.POST.get('registration_open')
        event_open = request.POST.get('event_open')
        event_over = request.POST.get('event_over')
        offers_certification = request.POST.get('offers_certification')
        print(title,description,location,poster,start_date,registration_open,event_open)
    
        if title:
            event.title = title
        if description:
            event.description =description
        if location:
            event.location = location
        if poster:
            event.poster = poster
        if start_date and start_time:
            combined_datetime = datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M")
            event.start_date = combined_datetime

        event.registration_open = registration_open
        event.event_open = event_open
        event.event_over = event_over
        event.offers_certification = offers_certification
        event.save()
        messages.success(request,'Changes applied sucessfully')
        return redirect('teacherEvent' , event.slug)
        
    return render(request, 'events/event_edit.html', {'event': event})

@login_required(login_url='login')
def add_coordinators(request,slug):
    users = User.objects.filter(groups__name__icontains='COORDINATOR').exclude(username = request.user.username)
    event = get_object_or_404(models.Event , slug=slug)
    if not is_event_host(request,event):
        return redirect('events')
    if request.method == "POST":
        user_id = request.POST.getlist('users')
        selected_users = User.objects.filter(id__in=user_id)
        try:
            for user in selected_users:
                models.EventCordinator.objects.create(event=event, user=user)
        except Exception as e:
            print(e)
            messages.error(request, "User is already Coordinator")
        return redirect('teacherEvent', event.slug)
    return render(request , 'events/add_coordinators.html',{'users':users})

@login_required(login_url='login')
def remove_coordinator(request,uid):
    coordinator = get_object_or_404(models.EventCordinator , uid=uid) 
    coordinator.delete()
    messages.success(request, "User removed from Coordinator")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def live_polling_qna(request):
    return render(request, 'events/live_poll_qna.html')

